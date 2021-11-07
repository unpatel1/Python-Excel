# 11/6/2021
# Automate Excel With Python - Python Excel Tutorial (OpenPyXL)
# https://www.youtube.com/watch?v=7YS6YDQKFh0

'''
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# instentiate a workbook (new or existing one)
wb = load_workbook('Grades.xlsx')

# accessing worksheets
ws = wb.active
print(ws)

# accessing cell values
print(ws['A1'].value)

# change cell value
ws['A2'].value = "Test" # ws['A2'] = "Test" also works
wb.save('Grades.xlsx')

# working with multiple sheets
print(wb.sheetnames)
    # access a sheet
print(wb['Sheet1'])
    # create a new sheet
wb.create_sheet("Test")
wb.save('Grades.xlsx')
print(wb.sheetnames)

# create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Data"

# insert data
ws.append(['Umesh', 'Is', 'Great', '!'])
ws.append(['Umesh', 'Is', 'Great', '!'])
ws.append(['Umesh', 'Is', 'Great', '!'])
ws.append(['Umesh', 'Is', 'Great', '!'])
ws.append(["end"])
wb.save('umesh.xlsx')

wb = load_workbook('umesh.xlsx')
ws = wb.active

for row in range(1, 11):
    for col in range(1, 5):
        char = get_column_letter(col)
#        print(ws[char + str(row)].value)
        ws[char + str(row)] = char + str(row)

# merge cells
ws.merge_cells("A1:D1")
ws.unmerge_cells("A1:D1")

# insert & delete rows
ws.insert_rows(7) # insert after row 7
ws.insert_rows(7)

ws.delete_rows(7)

# insert & delete columns
ws.insert_cols(2) # insert a column at column B - #2 column is B column

ws.delete_cols(2)

# copying and moving cells
ws.move_range("C1:D11", rows = 2, cols = 2)

wb.save('umesh.xlsx')

'''

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl import workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
	"Joe": {
		"math": 65,
		"science": 78,
		"english": 98,
		"gym": 89
	},
	"Bill": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Tim": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"Sally": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Jane": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}

wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ['Name'] + list(data['Joe'].keys())
ws.append(headings)

for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

# calculate average of all columns
for col in range(2, len(data['Joe']) + 2):
    char = get_column_letter(col)
    ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

# styling
for col in range(1, 6):
    ws[get_column_letter(col) + '1'].font = Font(bold = True, color = "0099CCFF")



wb.save("NewGrades.xlsx")




